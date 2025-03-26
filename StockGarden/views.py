from django.shortcuts import HttpResponse,render, get_object_or_404, redirect
from django.http import JsonResponse
from .models import *
from .forms import *
import logging
from django.core.paginator import Paginator
from django.contrib import messages  
from django.db.models import Sum,Q,F
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required



# Create your views here
logger = logging.getLogger(__name__)

@login_required
def home(request):
    try:
        total_sales = Sales.objects.aggregate(Sum('total'))['total__sum'] or 0 
        total_repair_cost = RepairDetail.objects.aggregate(total_repair_cost=Sum('repair_cost'))['total_repair_cost'] or 0
        total_amount = total_sales + total_repair_cost
        total_products = Product.objects.count()  
        pending_repairs = Repair.objects.filter(status='in-progress').count()  
        low_stock_count = Product.objects.filter(stock__lt=3).count() 

        context = {
            'total_amount': total_amount,
            'total_product': total_products,
            'pending_repairs': pending_repairs,
            'low_stock_count': low_stock_count 
        }
        return render(request, 'home.html', context)
    except Exception as e:
        logger.error(f"Error in home view: {e}")
        messages.error(request, 'An error occurred while loading the dashboard.')
    return render(request, '404.html', {"message": "An error occurred while loading the dashboard."})

@login_required
def get_chart_data(request):
    try:
        total_sales = Sales.objects.aggregate(total=Sum('total')).get('total', 0) or 0
        total_repair_cost = RepairDetail.objects.aggregate(total=Sum('repair_cost')).get('total', 0) or 0
        total_earnings = total_sales + total_repair_cost
            
        chart_data = {
            "sales": total_sales,
            "repair": total_repair_cost,
            "earnings": total_earnings
        }
        return JsonResponse(chart_data)
    except Exception as e:
        logger.error(f"Error in home view: {e}")
        messages.error(request, 'An error occurred while loading the chart.')
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def BrandList(request):
    try:
        query = request.GET.get('q', '')
        brands = Brand.objects.all().order_by('-id')
        if query:
            brands = brands.filter(name__icontains=query)
            
        for brand in brands:
            if not brand.image:
                brand.image_url = None
            else:
                brand.image_url = brand.image.url
                
        paginator = Paginator(brands, 10)
        page_number = request.GET.get('page') 
        page_obj = paginator.get_page(page_number)

        context = {
            "brand": page_obj,
            "query": query,
            }
        return render(request, 'brand.html', context) 
    except Exception as e:
        logger.error(f"Error in BrandListView: {e}")
        messages.error(request, 'An error occurred while loading the brand list.')
        return render(request, '404.html', {"message": "An error occurred while loading the brand list."})

@login_required
def BrandCreate(request, brand_id=None):
    try:
        if brand_id:
            brand = get_object_or_404(Brand, id=brand_id)
            form = BrandForm(request.POST or None, request.FILES or None, instance=brand)
            action = "Update"
        else:
            form = BrandForm(request.POST or None, request.FILES or None)
            action = "Create"

        if request.method == 'POST':
            if form.is_valid():
                form.save()
                messages.success(request, f'Brand {action.lower()}d successfully!')
                return redirect('brand')

        return render(request, 'brand_create.html', {'form': form, 'action': action})
    
    except Exception as e:
        messages.error(request, 'An error occurred. Please try again later.')
        return redirect('brand')
    
    except Exception as e:
        logger.error(f"Error in BrandCreateView: {e}")
        messages.error(request, 'An error occurred while processing the brand.')
        return render(request, '404.html', {"message": "An error occurred."})
    
@login_required
def BrandUpdate(request, pk):
    try:
        brand = get_object_or_404(Brand, pk=pk)

        if request.method == 'POST':
            form = BrandForm(request.POST or None, request.FILES or None, instance=brand)
            if form.is_valid():
                form.save()
                messages.success(request, f"The brand '{brand.name}' has been successfully updated.")
                return redirect('brand')
        else:
            form = BrandForm(request.POST or None, request.FILES or None, instance=brand)
        return render(request, 'brand_update.html', {'form': form, 'brand': brand})

    except Exception as e:
        logger.error(f"Error in BrandUpdateView for brand {pk}: {e}")
        messages.error(request, 'An error occurred while updating the brand.')
        return render(request, 'error.html', {"message": "An error occurred while updating the brand."})

@login_required
def BrandDelete(request, pk):
    try:
        brand = get_object_or_404(Brand, pk=pk)

        if request.method == 'POST':
            brand_name = brand.name  
            brand.delete()  
            messages.success(request, f"The brand '{brand_name}' has been successfully deleted.")
            return redirect('brand')  

        return render(request, 'brand_delete.html', {'brand': brand})
    except Exception as e:
        logger.error(f"Error in BrandDeleteView for brand {pk}: {e}")
        messages.error(request, 'An error occurred while deleting the brand.')
        return render(request, '404.html', {"message": "An error occurred while deleting the brand."})

@login_required
def CategoryList(request):
    try:
        query = request.GET.get('q', '').strip()
        category =Category.objects.all().order_by('-id')

        if query:
            logger.info(f"Searching for category: {query}")
            category = category.filter(
                Q(name__icontains=query) 
            )


        paginator =Paginator(category,10)
        page_number =request.GET.get('page')
        page_obj = paginator.get_page(page_number)        
        context={
            "category":page_obj,
            "query": query, 
        }
        return render(request,'category.html',context)
    except Exception as e:
        logger.error(f" Error in CategoryListView: {e}")
        return render(request, '404.html', {"message": "An error occurred."})
    
@login_required
def CategoryCreate(request,catagory_id=None):
    try:
        if catagory_id:
            category =get_object_or_404(Category,id=catagory_id)
            form = CategoryForm(request.POST or None,instance=category)
            action="Update"
        else:
            form = CategoryForm(request.POST or None)
            action="Create"
        if request.method == 'POST':
            if form.is_valid():
                form.save()
                messages.success(request,f'Caregory {action.lower()}d successfully!')
                return redirect('category')
        return render(request, 'category_create.html',{'form':form,'action':action})
    except Exception as e:
        logger.error(f"Error in Category Create View: {e}")
        messages.error(request, 'An error occurred while processing the category.')
        return render(request, '404.html', {"message": "An error occurred."})
    
@login_required    
def CategoryUpdate(request,pk):
    try:
        category=get_object_or_404(Category,pk=pk)
        if request.method == 'POST':
            form =CategoryForm(request.POST ,instance=category)
            if form.is_valid():
                form.save()
                messages.success(request,f'Category updated successfully!')
                return redirect('category')
        else:
            form =CategoryForm(instance=category)
        return render(request, 'category_update.html',{'form':form,'category':category})
    except Exception as e:
        logger.error(f"Error in CategoryUpdateView: {e}")
        messages.error(request, 'An error occurred while processing the category.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def CategoryDelete(request,pk):
    try:
        category=get_object_or_404(Category,pk=pk)
        if request.method == 'POST':
            category_name=category.name
            category.delete()
            messages.success(request,f'Category {category_name} deleted successfully!')
            return redirect('category')
        
        return render(request, 'category_delete.html',{'category':category})
    except Exception as e:
        logger.error(f"Error in CategoryDeleteView: {e}")
        messages.error(request, 'An error occurred while processing the category.')
        return render(request, '404.html', {"message": "An error occurred."})


@login_required
def PurchaseList(request):
    try:
        query = request.GET.get('q', '')
        purches=Purchase.objects.all()

        if query:
            purches = purches.filter(
                Q(vendor__company_name__icontains=query) |  
                Q(product__name__icontains=query) |       
                Q(description__icontains=query)      
            )

        pagination =Paginator(purches,10)
        page_number=request.GET.get('page')
        page_obj=pagination.get_page(page_number)
        context={
            'purches':page_obj,
        }
        return render(request, 'purchases.html',context)
    except Exception as e:
        logger.error(f"Error in PurchaseListView: {e}")
        messages.error(request, 'An error occurred while loading the purchase list.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def PurchaseCreate(request,pruchase_id=None):
    try:
        if pruchase_id:
            purchase=get_object_or_404(Purchase,id=pruchase_id)
            form=PurchaseForm(request.POST or None,instance=purchase)
            action='update'
        else:
            form=PurchaseForm(request.POST or None)
            action='create'
        if request.method == 'POST':
            if form.is_valid():
                form.save()
                messages.success(request,f'Purchase {action.lower()}d successfully!')
                return redirect('purchase')
        return render(request, 'purchases_create.html',{'form':form,'action':action})
    except Exception as e:
        logger.error(f"Error in PurchaseCreateView: {e}")
        messages.error(request, 'An error occurred while processing the purchase.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def PurchaseUpdate(request,pk):
    try:
        purchase = get_object_or_404(Purchase,pk=pk)
        if request.method == 'POST':
            form= PurchaseForm(request.POST,instance=purchase)
            if form.is_valid():
                form.save()
                messages.success(request,f'Purchase updated successfully!')
                return redirect('purchase')
        else:
            form = PurchaseForm(instance=purchase)
            return render(request, 'purchases_update.html',{'form':form,'purchase':purchase})
    except Exception as e:
        logger.error(request, 'An error occurred while processing the purchase.')
        messages.error(request, 'An error occurred while processing the purchase.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def PurchaseDelete(request,pk):
    try:
        purchase= get_object_or_404(Purchase,pk=pk)
        if request.method == 'POST':
            purchase_name=purchase.product.name
            purchase.delete()
            messages.success(request,f'Purchase {purchase_name} deleted successfully!')
            return redirect('purchase')
        return render(request, 'purchase_delete.html',{'purchase':purchase})
    except Exception as e:
        logger.error(f"Error in PurchaseDeleteView: {e}")
        messages.error(request, 'An error occurred while processing the purchase.')
        return render(request, '404.html', {"message": "An error occurred."})
    
        
@login_required
def ProductList(request):
    try:
        query = request.GET.get('q', '')
        product =Product.objects.all().order_by('-created_at')

        if query:
            product = product.filter(
                Q(name__icontains=query) |
                Q(description__icontains=query) |
                Q(brand__name__icontains=query) |
                Q(categories__name__icontains=query)
            )

        paginator =Paginator(product,10)
        page_number =request.GET.get('page')
        page_obj =paginator.get_page(page_number)
        
        context ={
            "products":page_obj,
            "query": query,
        }
        return render(request, 'product.html',context)

    except Exception as e:
        logger.error(f"Error in ProductListView: {e}")
        messages.error(request,"An error occurred while loading the product list.")
        return render(request, '404.html', {"message": "An error occurred."})
    
@login_required
def ProductUpdate(request, pk):
    try:
        product = get_object_or_404(Product, pk=pk)

        if request.method == 'POST':
            form = ProductForm(request.POST, request.FILES, instance=product)
            if form.is_valid():
                form.save()
                messages.success(request, 'Product updated successfully!')
                return redirect('product')
            else:
                messages.error(request, 'Please correct the errors below.')
        else:
            form = ProductForm(instance=product)

        return render(request, 'product_update.html', {'form': form, 'product': product})
    except Exception as e:
        logger.error(f"Error in ProductUpdateView: {e}")
        messages.error(request, 'An error occurred while updating the product.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def ProductDelete(request,pk):
    try:
        product = get_object_or_404(Product,pk=pk)
        if request.method == 'POST':
            product_name = product.name
            product.delete()
            messages.success(request,f'Product {product_name} deleted successfully!')
            return redirect('product')
        return render(request, 'product_delete.html',{'product':product})
    except Exception as e:
        logger.error(f"Error in ProductDeleteView: {e}")
        messages.error(request, 'An error occurred while processing the product.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def SalesList(request):
    try:
        query = request.GET.get('q', '')
        sales = Sales.objects.all().order_by('-id')

        if query:
            sales = sales.filter(
                Q(name__full_name__icontains=query) |
                Q(product__name__icontains=query) | 
                Q(contact_no__icontains=query) 
            )

        paginator = Paginator(sales, 10) 
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context={
            "sales":page_obj,
            "query": query,
        }
        return render(request, 'sales.html',context)
    except Exception as e:
        logger.error(f"Error in SalesListView: {e}")
        return render(request, '404.html', {"message": "An error occurred."})
    
@login_required
def SalesCreate(request,sales_id=None):
    try:
        if sales_id:
            sales = get_object_or_404(Sales,id = sales_id)
            form = SalesForm(request.POST or None,instance=sales)
            action = 'Update'
        else:
            form = SalesForm(request.POST or None)
            action ='Create'
        if request.method == 'POST':
            if form.is_valid():
                form.save()
                messages.success(request, f" Sales {action.lower()}d successfully!")
                return redirect('sales')
        return render(request, 'sales_create.html',{'form':form,'action':action})
    except Exception as e:
        logger.error(f"Error in SalesCreateView: {e}")
        messages.error(request, 'An error occurred while processing the sales.')
        return render(request, '404.html', {"message": "An error occurred."})
    
@login_required
def SalesUpdate(request, pk):
    try:
        sales = get_object_or_404(Sales, pk=pk)

        if request.method == 'POST': 
            form = SalesForm(request.POST, instance=sales) 
            if form.is_valid():
                form.save()
                messages.success(request, f'Sales updated successfully!')
                return redirect('sales') 
        else:
            form = SalesForm(instance=sales)

        return render(request, 'sales_update.html', {'form': form, 'sales': sales})
    except Exception as e:
        logger.error(f"Error in SalesUpdateView: {e}")
        messages.error(request, 'An error occurred while processing the sales.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def SalesDelete(request,pk):
    try:
        sales =get_object_or_404(Sales,pk=pk)
        if request.method == 'POST':
            sales_name = sales.name
            sales.delete()
            messages.success(request,f'Sales {sales_name} deleted successfully!')
            return redirect('sales')
        return render(request, 'sales_delete.html',{'sales':sales})
    except Exception as e:
        logger.error(f"Error in SalesDeleteView: {e}")
        messages.error(request, 'An error occurred while processing the sales.')
        return render(request, '404.html', {"message": "An error occurred."})



@login_required
def RepairList(request):
    try:
        query = request.GET.get('q', '')
        repair=Repair.objects.all().order_by('id')

        if query:
            repair = repair.filter(
                Q(product_name__icontains=query) | 
                Q(device_model__icontains=query) |
                Q(name__full_name__icontains=query) 
            )

        pagination=Paginator(repair,10)
        page_number=request.GET.get('page')
        page_obj=pagination.get_page(page_number)
        context={
            'repair':page_obj,
            'query': query,
        }
        return render(request, 'repair.html',context)
    except Exception as e:
        logger.error(f'An error occurred while loading the repair list:- {e}')
        messages.error(request, 'An error occurred while loading the repair list.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def RepairCreate(request,repair_id=None):
    try:
        if repair_id:
            repair=get_object_or_404(Repair,id=repair_id)
            form=RepairForm(request.POST or None,instance=repair)
            action='update'
        else:
            form=RepairForm(request.POST or None)
            action='create'
        if request.method == 'POST':
            if form.is_valid():
                form.save()
                messages.success(request,f'Repair {action.lower()}d successfully!')
                return redirect('repair')
            else:
                print("Form Errors:", form.errors) 
                messages.error(request, 'Form is invalid.')
        return render(request, 'repair_create.html',{'form':form,'action':action})
    except Exception as e:
        logger.error(f"Error in RepairCreateView: {e}")
        messages.error(request, 'An error occurred while processing the repair.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def RepairUpdate(request,pk):
    try:
        repair=get_object_or_404(Repair,pk=pk)
        if request.method =='POST':
            form=RepairForm(request.POST or None,instance=repair)
            if form.is_valid():
                form.save()
                messages.success(request,f'Repair updated successfully!')
                return redirect('repair')
        else:
            form=RepairForm(instance=repair)
            return render(request, 'repair_update.html',{'form':form,'repair':repair})
    except Exception as e:
        logger.error(f"Error in RepairUpdateView: {e}")
        messages.error(request, 'An error occurred while processing the repair.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def RepairDelete(request,pk):
    try:
        repair=get_object_or_404(Repair,pk=pk)
        if request.method == 'POST':
            repair_name=repair.name
            repair.delete()
            messages.success(request,f'Repair {repair_name} deleted successfully!')
            return redirect('repair')
        return render(request, 'repair_delete.html',{'repair':repair})
    except Exception as e:
        logger.error(f"Error in RepairDeleteView{pk}: {e}")
        messages.error(request, 'An error occurred while processing the repair.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def RepairDetailList(request):
    try:
        query = request.GET.get('q', '')
        repairdetail = RepairDetail.objects.all().order_by('id')

        if query:
            repairdetail = repairdetail.filter(
                Q(repair_order__product_name__icontains=query) |  # You can search by repair_order product_name
                Q(fixed_description__icontains=query) |  # Search by fixed_description
                Q(repair_action__icontains=query)  # Search by repair_action
            )
        pagination = Paginator(repairdetail, 10)
        page_number = request.GET.get('page') 
        page_obj = pagination.get_page(page_number)
        context = {
            'repairdetail': page_obj,
            'query': query,
        }
        return render(request, 'repairdetail.html', context)
    except Exception as e:
        logger.error(f"Error in RepairListView: {e}")
        messages.error(request, 'An error occurred while loading the repair detail list.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required    
def RepairDetailCreate(request,repairde_id=None):
    try:
        if repairde_id:
            repairdetail=get_object_or_404(RepairDetail,id=repairde_id)
            form=RepairDetail(request.POST or None,instance=repairdetail)
            action='Update'
        else:
            form = RepairDetailForm(request.POST or None)
            action = 'Create'
            if request.method == 'POST':
                if form.is_valid():
                    form.save()
                    messages.success(request,f'Repair Detail {action.lower()}d successfully!')
                    return redirect('repair_detail')
            return render(request, 'repair_detail_create.html',{'form':form,'action':action})
    except Exception as e:
        logger.error(f"Error in RepairDetailCreateView: {e}")
        messages.error(request, 'An error occurred while processing the repair detail.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def RepairDetailUpdate(request,pk):
    try:
        repairdetail=get_object_or_404(RepairDetail,pk=pk)
        if request.method == 'POST':
            form=RepairDetailForm(request.POST or None,instance=repairdetail)
            if form.is_valid():
                form.save()
                messages.success(request,f'Repair Detail updated successfully!')
                return redirect('repair_detail')
        else:
            form =RepairDetailForm(instance=repairdetail)
            return render(request, 'repair_detail_update.html', {'form': form, 'repairdetail': repairdetail})
    except Exception as e:
        logger.error(f"Error in RepairDetailUpdateView: {e}")
        messages.error(request, 'An error occurred while processing the repair detail.')
        return render(request, '404.html', {"message": "An error occurred."})
    
@login_required    
def RepairDetailDelete(request, pk):
    try:
        repairdetail = get_object_or_404(RepairDetail, pk=pk)

        if request.method == 'POST':
            repairdetail_name = repairdetail.repair_order.product_name
            repairdetail.delete()
            messages.success(request, f"Repair Detail '{repairdetail_name}' deleted successfully!")
            return redirect('repair_detail')  
        return render(request, 'repair_detail_delete.html', {'repairdetail': repairdetail, 'deleted': False})

    except Exception as e:
        logger.error(f"Error in RepairDetailDeleteView: {e}", exc_info=True)
        messages.error(request, 'An error occurred while processing the repair detail. Please try again later.')
        return render(request, '404.html', {"message": "An error occurred."})
    
@login_required
def ExpenseList(request):
    query = request.GET.get('search', '')
    expenses = Expense.objects.filter(
        description__icontains=query
    ) if query else Expense.objects.all()

    # Pagination
    paginator = Paginator(expenses, 10)  # 10 expenses per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'expense.html',{'expenses': page_obj, 'query': query})

@login_required
def ExpenseCreate(request,expense_id=None):
    try:
        if expense_id:
            expense=get_object_or_404(Expense,id=expense_id)
            form=ExpenseForm(request.POST or None,instance=expense)
            action='update'
        else:
            form=ExpenseForm(request.POST or None)
            action='create'
        if request.method == 'POST':
            if form.is_valid():
                form.save()
                messages.success(request,f'Expense {action.lower()}d successfully!')
                return redirect('expense')
            else:
                messages.error(request, 'Expense Form is invalid.')
        return render(request, 'expense_create.html',{'form':form,'action':action})
    except Exception as e:
        logger.error(f"Error in Expense Create: {e}")
        messages.error(request, 'An error occurred while processing the Expense Create.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def ExpenseUpdate(request,pk):
    try:
        expense=get_object_or_404(Expense,pk=pk)
        if request.method == 'POST':
            form=ExpenseForm(request.POST or None,instance=expense)
            if form.is_valid():
                form.save()
                messages.success(request,f'Expense updated successfully!')
                return redirect('expense')
        else:
            form =ExpenseForm(instance=expense)
            return render(request, 'expense_update.html', {'form': form, 'expenses': expense})
    except Exception as e:
        logger.error(f"Error in Expense Update: {e}")
        messages.error(request, 'An error occurred while processing the expense.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def ExpenseDelete(request,pk):
    try:
        expense= get_object_or_404(Expense,pk=pk)
        if request.method == 'POST':
            expense=expense.category.name
            expense.delete()
            messages.success(request,f'Expense {expense} deleted successfully!')
            return redirect('expense')
        return render(request, 'expense_delete.html',{'expenses':expense})
    except Exception as e:
        logger.error(f"Error in ExpenseDelete: {e}")
        messages.error(request, 'An error occurred while processing the Expense Delete.')
        return render(request, '404.html', {"message": "An error occurred Expense Delete."})

@login_required
def SalesInvoiceList(request):
    query = request.GET.get('search', '')
    invoices = SalesInvoice.objects.filter(
        invoice_number__icontains=query
        ) if query else SalesInvoice.objects.all()

        # Pagination
    paginator = Paginator(invoices, 10)  # 10 invoices per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = ({            
            'invoices': page_obj,
            'query': query
        })
    return render(request, 'sales_invoice.html', context)
    
@login_required
def SalesInvoiceUpdate(request, pk):
    try:
        # Fetch the single SalesInvoice object by primary key (pk)
        salesinvoice = get_object_or_404(SalesInvoice, pk=pk)
        
        if request.method == 'POST':
            form = SalesInvoiceForm(request.POST, instance=salesinvoice)
            if form.is_valid():
                form.save()
                messages.success(request, 'Sales Invoice updated successfully!')
                return redirect('salesinvoice')
        else:
            form = SalesInvoiceForm(instance=salesinvoice)
        
        # Pass the form and the single salesinvoice object to the template
        return render(request, 'sales_invoice_update.html', {'form': form, 'salesinvoice': salesinvoice})
    except Exception as e:
        logger.error(f"Error in Sales Invoice: {e}")
        messages.error(request, 'An error occurred while processing the Sales Invoice.')
        return render(request, '404.html', {"message": "An error occurred."})


@login_required
def RepairInvoiceList(request):
    query = request.GET.get('search', '')
    repair_invoices = RepairInvoice.objects.filter(
        invoice_number__icontains=query
    ) if query else RepairInvoice.objects.all()

    # Pagination
    paginator = Paginator(repair_invoices, 10)  # 10 invoices per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'invoices': page_obj,
        'query': query,
    }

    return render(request, 'repair_invoice.html', context)


    
    
@login_required
def RepairInvoiceUpdate(request, pk):
    try:
        repair_invoice = get_object_or_404(RepairInvoice, pk=pk)

        if request.method == 'POST':
            # Update the repair invoice using the form data
            form = RepairInvoiceForm(request.POST, instance=repair_invoice)
            if form.is_valid():
                form.save()
                messages.success(request, 'Repair Invoice updated successfully!')
                return redirect('repairinvoice')
        else:
            # Pre-fill the form with the existing data
            form = RepairInvoiceForm(instance=repair_invoice)

        return render(request, 'repair_invoice_update.html', {'form': form, 'repair_invoice': repair_invoice})

    except Exception as e:
        logger.error(f"Error in Repair Invoice: {e}")
        messages.error(request, 'An error occurred while processing the Repair Invoice.')
        return render(request, '404.html', {"message": "An error occurred."})


@login_required
def ReturnList(request):
    # Get search query from request
    query = request.GET.get('search', '')

    # Filter returns based on search query
    if query:
        returns = Return.objects.filter(
            Q(invoice__invoice_number__icontains=query) | 
            Q(product__name__icontains=query) | 
            Q(customer_name__icontains=query)
        )
    else:
        returns = Return.objects.all()

    # Paginate the returns
    paginator = Paginator(returns, 10)  # Show 10 returns per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Render the return list template
    return render(request, 'return.html', {
        'returns': page_obj,
        'query': query
    })

@login_required
def ReturnCreate(request,return_id=None):
    try:
        if return_id:
            returns=get_object_or_404(Return,id=return_id)
            form=ReturnForm(request.POST or None,instance=returns)
            action='update'
        else:
            form=ReturnForm(request.POST or None)
            action='create'
        if request.method == 'POST':
            if form.is_valid():
                form.save()
                messages.success(request,f'Return {action.lower()}d successfully!')
                return redirect('return')
            else:
                messages.error(request, 'Return Form is invalid.')
        return render(request, 'return_create.html',{'form':form,'action':action})
    except Exception as e:
        logger.error(f"Error in Return Create: {e}")
        messages.error(request, 'An error occurred while processing the Return Create.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def ReturnUpdate(request,pk):
    try:
        returns=get_object_or_404(Return,pk=pk)
        if request.method == 'POST':
            form=ReturnForm(request.POST or None,instance=returns)
            if form.is_valid():
                form.save()
                messages.success(request,f'Return updated successfully!')
                return redirect('return')
        else:
            form =ReturnForm(instance=returns)
            return render(request, 'return_update.html', {'form': form, 'returns': returns})
    except Exception as e:
        logger.error(f"Error in Return Update: {e}")
        messages.error(request, 'An error occurred while processing the Return.')
        return render(request, '404.html', {"message": "An error occurred."})


@login_required
def ReturnDelete(request,pk):
    try:
        returns= get_object_or_404(Return,pk=pk)
        if request.method == 'POST':
            returns=returns.invoice.invoice_number
            returns.delete()
            messages.success(request,f'Return {returns} deleted successfully!')
            return redirect('return')
        return render(request, 'return_delete.html',{'returns':returns})
    except Exception as e:
        logger.error(f"Error in Return Delete: {e}")
        messages.error(request, 'An error occurred while processing the Return Delete.')
        return render(request, '404.html', {"message": "An error occurred Return Delete."})

@login_required
def UserReportList(request):
    try:
        selected_role = request.GET.get('role', '') 
        users = User.objects.all().order_by('id')

        if selected_role:
            users = users.filter(role=selected_role)

        pagination = Paginator(users, 10)
        page_number = request.GET.get('page')
        page_obj = pagination.get_page(page_number)

        context = {
            'page_obj': page_obj,
            'selected_role': selected_role,
        }
        return render(request, 'user_report.html', context)

    except Exception as e:
        logger.error(f"Error in UserReportListView: {e}")
        messages.error(request, 'An error occurred while loading the report list.')
        return render(request, '404.html', {"message": "An error occurred."})

@login_required
def global_search(request):
    query = request.GET.get('q')
    context = {}

    try:
        if query:
            products = Product.objects.filter(Q(name__icontains=query) | Q(description__icontains=query))
            sales = Sales.objects.filter(Q(name__full_name__icontains=query) | Q(product__name__icontains=query))
            purchases = Purchase.objects.filter(Q(vendor__company_name__icontains=query) | Q(product__name__icontains=query))
            repairs = Repair.objects.filter(Q(product_name__icontains=query) | Q(device_model__icontains=query) | Q(name__full_name__icontains=query))
            brands = Brand.objects.filter(Q(name__icontains=query))
            categories = Category.objects.filter(Q(name__icontains=query))
            users = User.objects.filter(Q(full_name__icontains=query))

            context = {
                'products': products,
                'sales': sales,
                'purchases': purchases,
                'repairs': repairs,
                'brands': brands,
                'categories': categories,
                'users': users,
                'query': query,
            }
    except Exception as e:
        logger.error(f"Error occurred in global_search: {e}", exc_info=True)
        context['error'] = "An error occurred while processing your search. Please try again."

    return render(request, 'search_results.html', context)

@login_required
def SalesReportList(request):
    context = {}

    try:
        sales = Sales.objects.annotate(total_amount=F('quantity') * F('price'))

        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        if start_date and end_date:
            sales = sales.filter(created_at__date__range=[start_date, end_date])

        total_quantity = sales.aggregate(Sum('quantity'))['quantity__sum'] or 0

        context = { 
            'sales': sales,
            'total_quantity': total_quantity,
        }
    
    except Exception as e:
        logger.error(f"Error occurred in SalesReportListView: {e}", exc_info=True)
        context['error'] = "An error occurred while generating the sales report."

    return render(request, 'sales_report.html', context)

@login_required
def StockReportList(request):
    # Get query parameters for filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    product_name = request.GET.get('product_name')
    low_stock_threshold = request.GET.get('low_stock_threshold')
    total_sales_threshold = request.GET.get('total_sales_threshold')

    # Filter the products based on the provided parameters
    products = Product.objects.all()

    if start_date:
        products = products.filter(created_at__gte=start_date)
    if end_date:
        products = products.filter(created_at__lte=end_date)
    if product_name:
        products = products.filter(name__icontains=product_name)
    if low_stock_threshold:
        products = products.filter(stock__lt=low_stock_threshold)
    if total_sales_threshold:
        products = products.annotate(
            total_sales=Sum('sales__quantity') 
        ).filter(total_sales__gte=total_sales_threshold)

    sales_data = Sales.objects.filter(product__in=products).values('product').annotate(total_sales=Sum('quantity'))
    purchases_data = Purchase.objects.filter(product__in=products).values('product').annotate(total_purchased=Sum('quantity'))

    sales_dict = {sale['product']: sale['total_sales'] for sale in sales_data}
    purchases_dict = {purchase['product']: purchase['total_purchased'] for purchase in purchases_data}

    for product in products:
        product.total_sales = sales_dict.get(product.id, 0)
        product.total_purchased = purchases_dict.get(product.id, 0)

    context = {
        'products': products,
        'start_date': start_date,
        'end_date': end_date,
        'product_name': product_name,
        'low_stock_threshold': low_stock_threshold,
        'total_sales_threshold': total_sales_threshold,
    }

    return render(request, 'stock_report.html', context)


def generate_pdf(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    product_name = request.GET.get("product_name")
    
    products = Product.objects.all()
    if product_name:
        products = products.filter(name__icontains=product_name)
    
    sales = Sales.objects.all()
    purchases = Purchase.objects.all()

    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            if start_date <= end_date:
                sales = sales.filter(created_at__date__range=[start_date, end_date])  
                purchases = purchases.filter(created_at__date__range=[start_date, end_date])  
        except ValueError:
            pass

    total_sells = sum(sale.quantity for sale in sales if sale.quantity is not None)
    total_purchase = sum(purchase.quantity for purchase in purchases if purchase.quantity is not None)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(name="Title", fontSize=18, textColor=colors.darkblue, alignment=1, spaceAfter=15)
    subtitle_style = ParagraphStyle(name="Subtitle", fontSize=14, textColor=colors.darkred, spaceAfter=10)
    normal_style = ParagraphStyle(name="Normal", fontSize=12, spaceAfter=8)

    elements = []

    # Header
    elements.append(Paragraph("Sales and Purchase Report", title_style))
    elements.append(Spacer(1, 8))

    # Date Range
    date_range_text = f"Date Range: {start_date} to {end_date}" if start_date and end_date else "Date Range: All Time"
    elements.append(Paragraph(date_range_text, subtitle_style))
    elements.append(Spacer(1, 8))

    # Product Filter
    if product_name:
        elements.append(Paragraph(f"Product: {product_name}", normal_style))
        elements.append(Spacer(1, 8))

    # Summary Section
    summary_data = [
        ["Total Sales", total_sells],
        ["Total Purchases", total_purchase],
    ]
    summary_table = Table(summary_data, colWidths=[150, 200])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 15))

    # Sales Table
    sales_data = [['Sale Date', 'Product', 'Quantity', 'Price']]
    for sale in sales:
        sales_data.append([
            sale.created_at.strftime('%Y-%m-%d'),
            sale.product.name,
            sale.quantity,
            f"${sale.price:.2f}",
        ])

    sales_table = Table(sales_data, colWidths=[100, 200, 100, 100])
    sales_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(Paragraph("Sales Data", subtitle_style))
    elements.append(sales_table)
    elements.append(Spacer(1, 15))

    # Purchase Table
    purchase_data = [['Purchase Date', 'Product', 'Vendor', 'Quantity', 'Price', 'Total Value']]
    for purchase in purchases:
        purchase_data.append([
            purchase.created_at.strftime('%Y-%m-%d'),
            purchase.product.name if purchase.product else "N/A",
            purchase.vendor.name,
            purchase.quantity,
            f"${purchase.price:.2f}",
            f"${purchase.total_value:.2f}",
        ])

    purchase_table = Table(purchase_data, colWidths=[100, 200, 150, 100, 100, 100])
    purchase_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkred),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(Paragraph("Purchase Data", subtitle_style))
    elements.append(purchase_table)
    elements.append(Spacer(1, 15))

    # Footer (Page Numbers)
    def add_footer(canvas, doc):
        canvas.setFont("Helvetica", 9)
        canvas.drawString(500, 20, f"Page {doc.page}")

    # Build the PDF
    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)

    return response

@login_required
def generate_excel(request):
    try:
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        product_name = request.GET.get("product_name")
        
        # Initialize queryset with select_related to optimize database queries
        products = Product.objects.select_related('categories').all()
        if product_name:
            products = products.filter(name__icontains=product_name)
        
        # Optimize queries with select_related
        sales = Sales.objects.select_related('product')
        purchases = Purchase.objects.select_related('product')

        # Improved date validation and parsing
        if start_date and end_date:
            # Validate date format
            date_format = "%Y-%m-%d"
            try:
                # Parse dates and convert to date objects
                parsed_start_date = datetime.strptime(start_date, date_format).date()
                parsed_end_date = datetime.strptime(end_date, date_format).date()
                
                # Validate date range
                if parsed_start_date > parsed_end_date:
                    messages.error(request, "Start date must be before or equal to end date")
                    return redirect('stock_report')
                
                # Apply date filters
                sales = sales.filter(created_at__date__range=[parsed_start_date, parsed_end_date])
                purchases = purchases.filter(created_at__date__range=[parsed_start_date, parsed_end_date])
                
            except ValueError:
                messages.error(request, "Invalid date format. Please use YYYY-MM-DD format (e.g., 2024-01-31)")
                return redirect('stock_report')

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Report"

        # Add headers with styling
        headers = ["ID", "Product Name", "Category", "Stock", "Sold", "Purchased", "Date"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Add product data
        for product in products:
            try:
                # Calculate totals
                total_sold = (
                    sales.filter(product=product)
                    .aggregate(total_sold=models.Sum('quantity'))
                    ['total_sold'] or 0
                )

                total_purchased = (
                    purchases.filter(product=product)
                    .aggregate(total_purchased=models.Sum('quantity'))
                    ['total_purchased'] or 0
                )

                # Add row data
                ws.append([
                    product.id,
                    product.name,
                    product.categories.name if product.categories else "No Category",
                    product.stock,
                    total_sold,
                    total_purchased,
                    # Assuming product has a created_at field (you can adjust it to the correct field name if needed)
                    product.created_at.strftime("%Y-%m-%d") if product.created_at else "N/A"
                ])
            except Exception as row_error:
                logger.error(f"Error processing product {product.id}: {str(row_error)}")
                continue

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Generate response with current date in filename
        current_date = datetime.now().strftime("%Y%m%d")
        filename = f"stock_report_{current_date}.xlsx"
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        messages.error(request, "An error occurred while generating the Excel report")
        return redirect('stock_report')



@login_required
def RepairReportList(request):
    try:
        status = request.GET.get('status', '')  
        customer_id = request.GET.get('customer', '')  

        repairs = Repair.objects.all().order_by('-created_at')  

        if status:
            repairs = repairs.filter(status=status)
        if customer_id:
            repairs = repairs.filter(name_id=customer_id) 

        paginator = Paginator(repairs, 10)
        page_number = request.GET.get('page')

        try:
            repairs_page = paginator.get_page(page_number)
        except Exception as e:
            logger.warning(f"Pagination error: {e}")
            repairs_page = paginator.get_page(1)  # Default to first page if error occurs

        customers = User.objects.filter(role="Customer").order_by('full_name')

        context = {
            'repairs': repairs_page,
            'customers': customers,
            'selected_status': status,
            'selected_customer': customer_id,
        }

        logger.info("Repair report generated successfully.")

    except Exception as e:
        logger.error(f"Error in RepairReportListView: {e}", exc_info=True)
        context = {"error": "An error occurred while generating the repair report."}
    return render(request, 'repair_report.html', context)

@login_required
def RepairDetailReportList(request):
    try:
        repair_order_id = request.GET.get('repair_order', '')  
        repair_action = request.GET.get('repair_action', '')  

        repair_details = RepairDetail.objects.select_related('repair_order').order_by('-created_at')

        if repair_order_id:
            repair_details = repair_details.filter(repair_order_id=repair_order_id)
        if repair_action:
            repair_details = repair_details.filter(repair_action=repair_action)

        paginator = Paginator(repair_details, 10)
        page_number = request.GET.get('page')

        try:
            repair_details_page = paginator.get_page(page_number)
        except Exception as e:
            logger.warning(f"Pagination error in RepairDetailReportListView: {e}")
            repair_details_page = paginator.get_page(1)  

        repair_orders = Repair.objects.all().order_by('-created_at')

        context = {
            'repair_details': repair_details_page,
            'repair_orders': repair_orders,
            'selected_repair_order': repair_order_id,
            'selected_repair_action': repair_action,
        }

        logger.info("Repair detail report generated successfully.")

    except Exception as e:
        logger.error(f"Error in RepairDetailReportListView: {e}", exc_info=True)
        context = {"error": "An error occurred while generating the repair detail report."}

    return render(request, 'repair_detail_report.html', context)

@login_required
@csrf_exempt
def get_product_price(request):
    try:
        product_id = request.GET.get("product_id")

        if not product_id:
            logger.warning("Invalid request: Missing product_id")
            return JsonResponse({"error": "Invalid request"}, status=400)

        try:
            product = Product.objects.get(id=product_id)
            logger.info(f"Product found: {product.name} (ID: {product_id}) - Price: {product.price}")
            return JsonResponse({"price": product.price})

        except Product.DoesNotExist:
            logger.error(f"Product with ID {product_id} not found")
            return JsonResponse({"error": "Product not found"}, status=404)

    except Exception as e:
        logger.critical(f"Unexpected error in get_product_price: {e}", exc_info=True)
        return JsonResponse({"error": "An internal error occurred"}, status=500)



